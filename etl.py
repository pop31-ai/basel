"""ETL: извлечение данных из внешних файлов (банковские выписки, CSV и т.п.),
преобразование в проводки по правилам, загрузка в журнал.

Конфиг (JSON):
{
  "name": "bank_alfa",
  "source": {"type": "csv", "path": "in/alfa.csv", "delimiter": ";",
             "encoding": "utf-8", "date_format": "%d.%m.%Y"},
  "columns": {"date": "Дата", "desc": "Назначение платежа", "amount": "Сумма",
              "ref": "№ док", "counterparty": "Контрагент"},
  "chart": "ru",                      // опционально: загрузить систему счетов
  "rules": [
     {"when": {"counterparty": "ИП Иванов"}, "debit": "51", "credit": "62",
      "amount": "{amount}", "analytics": {"buyer": "ИП Иванов"}},
     {"when": {"desc": ".*зарплат.*"}, "debit": "70", "credit": "51",
      "amount": "{amount}"},
     {"always": true, "skip": true}   // прочее — пропустить
  ],
  "dedup_key": "ref"                  // ключ для дедупликации
}
"""
from __future__ import annotations
import csv, io, json, re
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional

from .core import Line, money
from .charts import Chart, builtin
from .storage import Store
from .journal import Journal
from .numbering import Numbering


# ---------- извлечение ----------
class ETLSource:
    """Читает исходный файл в список записей (dict)."""

    def __init__(self, cfg: Dict):
        self.cfg = cfg["source"]
        self.columns = cfg.get("columns", {})

    def extract(self, path: str = None) -> List[Dict]:
        p = Path(path or self.cfg.get("path"))
        if not p.exists():
            raise FileNotFoundError(f"Файл не найден: {p}")
        t = self.cfg.get("type", p.suffix.lower().lstrip("."))
        if t == "csv":
            return self._csv(p)
        if t in ("json", "jsonl", "ndjson"):
            return self._json(p)
        if t in ("xlsx", "xls"):
            return self._xlsx(p)
        raise ValueError(f"Неизвестный тип источника: {t}")

    @staticmethod
    def _row_filter(row: Dict) -> Dict:
        return {k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()}

    def _csv(self, p: Path) -> List[Dict]:
        delim = self.cfg.get("delimiter") or ";"
        with open(p, "r", encoding=self.cfg.get("encoding", "utf-8-sig"),
                  newline="") as f:
            reader = csv.DictReader(f, delimiter=delim)
            return [self._row_filter(r) for r in reader]

    def _json(self, p: Path) -> List[Dict]:
        data = json.loads(p.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            for key in ("rows", "data", "transactions"):
                if key in data:
                    return [self._row_filter(r) for r in data[key]]
            raise ValueError("JSON-источник должен содержать массив или ключ rows/data/transactions")
        return [self._row_filter(r) for r in data]

    def _xlsx(self, p: Path) -> List[Dict]:
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError("Для XLSX установите openpyxl: pip install openpyxl") from e
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb[self.cfg.get("sheet", wb.sheetnames[0])]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(it)]
        rows = []
        for vals in it:
            rows.append({header[i]: vals[i] for i in range(min(len(header), len(vals)))
                         if header[i]})
        wb.close()
        return rows


# ---------- преобразование ----------
def _parse_amount(s):
    from decimal import Decimal
    if isinstance(s, (int, float, Decimal)):
        return money(s)
    t = str(s).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", t)
    if not m:
        return money(0)
    return money(Decimal(m.group()))


class ETLTransform:
    """Преобразует записи в строки проводок по правилам."""

    def __init__(self, cfg: Dict, chart: Chart, journal: Journal):
        self.cfg = cfg
        self.columns = cfg.get("columns", {})
        self.rules = cfg.get("rules", [])
        self.chart = chart
        self.journal = journal
        self.date_format = cfg.get("source", {}).get("date_format", "%Y-%m-%d")

    def run(self, rows: List[Dict], start_idx: int = 0) -> List[Dict]:
        result = []
        for i, raw in enumerate(rows):
            row = {k: v for k, v in raw.items()}
            action = self._match(row)
            if action is None or action.get("skip"):
                continue
            entry = self._build(row, action, i + start_idx)
            result.append(entry)
        return result

    def _match(self, row: Dict) -> Optional[Dict]:
        for rule in self.rules:
            if rule.get("always"):
                return rule
            when = rule.get("when", {})
            matched = True
            for field, pattern in when.items():
                val = str(self._col(row, field) or "")
                if not re.fullmatch(pattern, val):
                    matched = False
                    break
            if matched:
                return rule
        return None

    def _col(self, row: Dict, logical: str):
        return row.get(logical, row.get(self.columns.get(logical, logical), ""))

    def _build(self, row: Dict, action: Dict, idx: int) -> Dict:
        amt = abs(_parse_amount(self._col(row, "amount")))
        d = self._parse_date(self._col(row, "date"))
        desc = str(self._col(row, "desc") or self._col(row, "counterparty") or "ETL")
        ref = self._col(row, "ref")
        analytics = {}
        for k, v in action.get("analytics", {}).items():
            analytics[k] = v
        # аналитика из колонок (подстановка {col})
        for k in ("counterparty", "buyer"):
            val = self._col(row, k)
            if val:
                analytics[k] = val
        lines = []
        if "debit" in action and "credit" in action:
            debit = action["debit"]
            credit = action["credit"]
            amount_expr = action.get("amount", "{amount}")
            amt_v = _parse_amount(self._subst(amount_expr, row, amt))
            lines.append(Line(debit, credit, amt_v, dict(analytics)))
        elif "template" in action:
            from .templates import Templates, substitute
            tmpl = Templates(self.journal.store).get(action["template"])
            params = {k: substitute(str(v), row) for k, v in action.get("params", {}).items()}
            params["sum"] = str(amt)
            for ln in tmpl.lines:
                from .templates import eval_amount
                amt_l = eval_amount(ln["amount"], params)
                lines.append(Line(ln["debit"], ln["credit"], amt_l, dict(analytics)))
        if not lines:
            return {"skip": True}
        no = None  # номер назначается при загрузке
        return {"date": d, "desc": desc, "ref": ref, "lines": lines, "idx": idx}

    def _subst(self, s: str, row: Dict, amt) -> str:
        out = str(s)
        for key, val in [("amount", str(amt))] + [(k, str(v)) for k, v in row.items()]:
            out = out.replace("{" + key + "}", str(val))
        return out

    def _parse_date(self, s) -> date:
        if isinstance(s, datetime):
            return s.date()
        if isinstance(s, date):
            return s
        s = str(s).strip()
        for fmt in (self.date_format, "%Y-%m-%d", "%d.%m.%Y", "%Y%m%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Некорректная дата: {s!r}")


# ---------- загрузка ----------
class ETLLoader:
    """Загружает преобразованные записи в журнал с дедупликацией."""

    def __init__(self, journal: Journal, numbering: Numbering, cfg: Dict):
        self.journal = journal
        self.numbering = numbering
        self.cfg = cfg
        self.dedup_key = cfg.get("dedup_key")

    def dedup_refs(self) -> set:
        if not self.dedup_key:
            return set()
        refs = set()
        for e in self.journal.entries():
            for l in e.lines:
                if self.dedup_key in l.analytics:
                    refs.add(str(l.analytics[self.dedup_key]))
        return refs

    def load(self, plans: List[Dict], dry_run: bool = False) -> Dict:
        existing = self.dedup_refs()
        loaded, skipped, errors = 0, 0, []
        for p in plans:
            if p.get("skip"):
                continue
            if self.dedup_key:
                ref = str(p.get("ref") or "")
                if ref and ref in existing:
                    skipped += 1
                    continue
            lines = p["lines"]
            no = self.numbering.next("entry", p["date"])
            desc = p["desc"]
            ref_an = {self.dedup_key: str(p["ref"])} if self.dedup_key and p.get("ref") else {}
            for l in lines:
                l.analytics.update(ref_an)
            try:
                if not dry_run:
                    self.journal.post(p["date"], lines, desc=desc, no=no,
                                      source=f"etl:{self.cfg.get('name', 'etl')}")
                loaded += 1
                if self.dedup_key:
                    existing.add(str(p["ref"]))
            except Exception as e:
                errors.append({"idx": p.get("idx"), "error": str(e)})
        if not dry_run:
            self.journal.store.write_json("etl_state.json",
                                          {"source": self.cfg.get("name"),
                                           "loaded": loaded, "skipped": skipped,
                                           "at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        return {"loaded": loaded, "skipped": skipped, "errors": errors}


class ETL:
    """Полный конвейер: extract -> transform -> load."""

    def __init__(self, store: Store, cfg_path: str):
        self.store = store
        self.cfg = json.loads(Path(cfg_path).read_text(encoding="utf-8"))
        self.name = self.cfg.get("name", "etl")
        # система счетов: из конфига или из базы
        chart_name = self.cfg.get("chart")
        if chart_name:
            chart = builtin(chart_name)
            self.store.write_json("chart.json", chart.to_dict())
        else:
            chart = Chart.from_dict(self.store.read_json("chart.json"))
        self.chart = chart
        self.journal = Journal(self.store, chart)
        self.numbering = Numbering(self.store)

    def preview(self, path: str = None, limit: int = 20) -> List[Dict]:
        rows = ETLSource(self.cfg).extract(path)
        plans = ETLTransform(self.cfg, self.chart, self.journal).run(rows)
        out = []
        for p in plans[:limit]:
            if p.get("skip"):
                continue
            out.append({"date": str(p["date"]), "desc": p["desc"],
                        "lines": [l.to_dict() for l in p["lines"]], "ref": p.get("ref")})
        return out

    def run(self, path: str = None, dry_run: bool = False) -> Dict:
        rows = ETLSource(self.cfg).extract(path)
        plans = ETLTransform(self.cfg, self.chart, self.journal).run(rows)
        res = ETLLoader(self.journal, self.numbering, self.cfg).load(plans, dry_run=dry_run)
        res["n_source"] = len(rows)
        return res